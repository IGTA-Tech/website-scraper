"""
Data processing and export module
Generates beautiful Excel and CSV reports from scraped data
"""
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from src.config import settings

console = Console()


class DataProcessor:
    """
    Process and export scraped data to Excel/CSV with beautiful formatting
    """

    def __init__(self, output_dir: Optional[str] = None):
        """
        Initialize data processor

        Args:
            output_dir: Directory for output files
        """
        self.output_dir = Path(output_dir or settings.output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _clean_data(self, data: List[Dict]) -> pd.DataFrame:
        """
        Clean and structure data for export

        Args:
            data: List of page data dictionaries

        Returns:
            Pandas DataFrame
        """
        # Define column order and labels
        columns = {
            "url": "URL",
            "title": "Title",
            "meta_description": "Meta Description",
            "meta_keywords": "Meta Keywords",
            "status_code": "Status Code",
            "word_count": "Word Count",
            "h1_count": "H1 Count",
            "h1_tags": "H1 Tags",
            "internal_links_count": "Internal Links",
            "external_links_count": "External Links",
            "image_count": "Images",
            "ai_summary": "AI Summary",
            "ai_topic": "Primary Topic",
            "ai_keywords": "AI Keywords",
            "ai_audience": "Target Audience",
            "ai_quality_score": "Quality Score",
            "ai_seo_score": "SEO Score",
            "ai_content_type": "Content Type",
            "ai_sentiment": "Sentiment",
        }

        # Extract relevant fields
        processed_data = []
        for item in data:
            row = {}
            for key, label in columns.items():
                value = item.get(key, "")

                # Handle lists (H1 tags, keywords)
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)

                row[label] = value

            processed_data.append(row)

        return pd.DataFrame(processed_data)

    def _format_excel(self, filepath: Path):
        """
        Apply beautiful formatting to Excel file

        Args:
            filepath: Path to Excel file
        """
        wb = load_workbook(filepath)
        ws = wb.active

        # Define styles
        header_fill = PatternFill(start_color="00D9FF", end_color="00D9FF", fill_type="solid")
        header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Inter", size=10)
        border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Format data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = cell_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        # Adjust column widths
        column_widths = {
            "A": 50,  # URL
            "B": 40,  # Title
            "C": 50,  # Meta Description
            "D": 30,  # Meta Keywords
            "E": 12,  # Status Code
            "F": 12,  # Word Count
            "G": 10,  # H1 Count
            "H": 30,  # H1 Tags
            "I": 12,  # Internal Links
            "J": 12,  # External Links
            "K": 10,  # Images
            "L": 60,  # AI Summary
            "M": 20,  # Primary Topic
            "N": 40,  # AI Keywords
            "O": 30,  # Target Audience
            "P": 12,  # Quality Score
            "Q": 12,  # SEO Score
            "R": 20,  # Content Type
            "S": 15,  # Sentiment
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add filters
        ws.auto_filter.ref = ws.dimensions

        # Save formatted workbook
        wb.save(filepath)

    def export_to_excel(
        self,
        data: List[Dict],
        filename: Optional[str] = None,
        format_file: bool = True
    ) -> Path:
        """
        Export data to Excel file

        Args:
            data: List of page data dictionaries
            filename: Custom filename (auto-generated if None)
            format_file: Whether to apply formatting

        Returns:
            Path to created file
        """
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            domain = data[0]['url'].split('/')[2].replace('.', '_') if data else 'scrape'
            filename = f"{domain}_scrape_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Create DataFrame
        df = self._clean_data(data)

        # Export to Excel
        console.print(f"\n[cyan]📊 Exporting to Excel...[/cyan]")
        df.to_excel(filepath, index=False, sheet_name="Scrape Results")

        # Apply formatting
        if format_file:
            console.print(f"[cyan]✨ Applying formatting...[/cyan]")
            self._format_excel(filepath)

        console.print(f"[green]✓ Excel file created: {filepath}[/green]")

        return filepath

    def export_to_csv(
        self,
        data: List[Dict],
        filename: Optional[str] = None
    ) -> Path:
        """
        Export data to CSV file

        Args:
            data: List of page data dictionaries
            filename: Custom filename (auto-generated if None)

        Returns:
            Path to created file
        """
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            domain = data[0]['url'].split('/')[2].replace('.', '_') if data else 'scrape'
            filename = f"{domain}_scrape_{timestamp}.csv"

        filepath = self.output_dir / filename

        # Create DataFrame
        df = self._clean_data(data)

        # Export to CSV
        console.print(f"\n[cyan]📄 Exporting to CSV...[/cyan]")
        df.to_csv(filepath, index=False, encoding='utf-8')

        console.print(f"[green]✓ CSV file created: {filepath}[/green]")

        return filepath

    def generate_summary(self, data: List[Dict], stats: Dict) -> Dict:
        """
        Generate summary statistics from scraped data

        Args:
            data: List of page data
            stats: Scraping statistics

        Returns:
            Dictionary of summary statistics
        """
        df = self._clean_data(data)

        # Calculate averages and totals
        summary = {
            "total_pages": len(data),
            "avg_word_count": df["Word Count"].mean() if "Word Count" in df.columns else 0,
            "avg_quality_score": df["Quality Score"].mean() if "Quality Score" in df.columns else 0,
            "avg_seo_score": df["SEO Score"].mean() if "SEO Score" in df.columns else 0,
            "total_internal_links": df["Internal Links"].sum() if "Internal Links" in df.columns else 0,
            "total_external_links": df["External Links"].sum() if "External Links" in df.columns else 0,
            "total_images": df["Images"].sum() if "Images" in df.columns else 0,
        }

        # Top topics
        if "Primary Topic" in df.columns:
            top_topics = df["Primary Topic"].value_counts().head(5).to_dict()
            summary["top_topics"] = top_topics

        # Top content types
        if "Content Type" in df.columns:
            content_types = df["Content Type"].value_counts().to_dict()
            summary["content_types"] = content_types

        # Sentiment distribution
        if "Sentiment" in df.columns:
            sentiment_dist = df["Sentiment"].value_counts().to_dict()
            summary["sentiment_distribution"] = sentiment_dist

        # Top performing pages
        if "Quality Score" in df.columns:
            top_pages = df.nlargest(5, "Quality Score")[["URL", "Title", "Quality Score"]].to_dict('records')
            summary["top_pages"] = top_pages

        # Add scraping stats
        summary.update(stats)

        return summary

    def print_summary_table(self, summary: Dict):
        """
        Print beautiful summary table

        Args:
            summary: Summary statistics dictionary
        """
        from rich.table import Table
        from rich.panel import Panel

        # Create main summary table
        table = Table(title="📊 Scrape Summary Report", show_header=True)
        table.add_column("Metric", style="cyan", width=30)
        table.add_column("Value", style="green", width=20)

        table.add_row("Total Pages Scraped", str(summary.get("total_pages", 0)))
        table.add_row("Avg Word Count", f"{summary.get('avg_word_count', 0):.0f}")
        table.add_row("Avg Quality Score", f"{summary.get('avg_quality_score', 0):.1f}/10")
        table.add_row("Avg SEO Score", f"{summary.get('avg_seo_score', 0):.1f}/10")
        table.add_row("Total Images", str(summary.get("total_images", 0)))
        table.add_row("Processing Time", f"{summary.get('duration', 0):.2f}s")

        console.print("\n")
        console.print(table)

        # Top topics
        if "top_topics" in summary:
            topics_table = Table(title="🏆 Top Topics", show_header=True)
            topics_table.add_column("Topic", style="cyan")
            topics_table.add_column("Count", style="green")

            for topic, count in summary["top_topics"].items():
                topics_table.add_row(topic, str(count))

            console.print("\n")
            console.print(topics_table)

        # Top pages
        if "top_pages" in summary:
            pages_table = Table(title="⭐ Top Performing Pages", show_header=True)
            pages_table.add_column("Title", style="cyan", width=40)
            pages_table.add_column("Score", style="green", width=10)

            for page in summary["top_pages"]:
                title = page.get("Title", "")[:40]
                score = page.get("Quality Score", 0)
                pages_table.add_row(title, f"{score}/10")

            console.print("\n")
            console.print(pages_table)

        console.print("\n")


# Convenience functions
def export_to_excel(data: List[Dict], filename: Optional[str] = None) -> Path:
    """Export data to Excel"""
    processor = DataProcessor()
    return processor.export_to_excel(data, filename)


def export_to_csv(data: List[Dict], filename: Optional[str] = None) -> Path:
    """Export data to CSV"""
    processor = DataProcessor()
    return processor.export_to_csv(data, filename)
